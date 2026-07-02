"""Convert a filled-in tier2_gold_labeling_worksheet.xlsx into
eval/gold/analysis_gold.jsonl (the format eval/tier2_analysis.py expects).

Expects the 'Labeling' sheet with columns:
    review_id | stars (ref only) | review_text | sentiment | aspect: food_quality
    | aspect: staff_attitude | aspect: pricing | aspect: wait_time
    | aspect: ambience | aspect: cleanliness | aspect: other

A row is skipped (with a warning) if 'sentiment' is blank - i.e. not yet labeled.
An aspect column counts as "mentioned" only if it has a non-blank value.

Run (from backend/):
    python eval/gold/build_gold_jsonl.py <filled_xlsx_path> [--out eval/gold/analysis_gold.jsonl]
"""

from __future__ import annotations

import argparse
import json
import sys

from openpyxl import load_workbook

ASPECT_COLUMNS = {
    5: "food_quality",
    6: "staff_attitude",
    7: "pricing",
    8: "wait_time",
    9: "ambience",
    10: "cleanliness",
    11: "other",
}
VALID_SENTIMENTS = {"positive", "negative", "neutral", "mixed"}
VALID_ASPECT_LABELS = {"positive", "negative", "neutral"}


def build(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Labeling"]

    rows: list[dict] = []
    skipped = 0
    for r in range(2, ws.max_row + 1):
        review_id = ws.cell(row=r, column=1).value
        if not review_id:
            continue
        sentiment = (ws.cell(row=r, column=4).value or "").strip().lower() or None

        if sentiment is None:
            skipped += 1
            print(f"warning: row {r} ({review_id}) has no sentiment - skipped (not yet labeled)")
            continue
        if sentiment not in VALID_SENTIMENTS:
            raise ValueError(f"row {r} ({review_id}): sentiment '{sentiment}' not in {VALID_SENTIMENTS}")

        aspects = []
        for col, category in ASPECT_COLUMNS.items():
            raw = ws.cell(row=r, column=col).value
            if raw is None or str(raw).strip() == "":
                continue
            label = str(raw).strip().lower()
            if label not in VALID_ASPECT_LABELS:
                raise ValueError(f"row {r} ({review_id}): aspect '{category}' label '{label}' not in {VALID_ASPECT_LABELS}")
            aspects.append({"category": category, "label": label})

        rows.append({"review_id": review_id, "sentiment": sentiment, "aspects": aspects})

    print(f"built {len(rows)} labeled rows ({skipped} skipped as unlabeled)")
    return rows


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("--out", default="eval/gold/analysis_gold.jsonl")
    args = ap.parse_args(argv[1:])

    rows = build(args.xlsx_path)
    if not rows:
        print("no labeled rows found - nothing written")
        return 1
    if len(rows) < 30:
        print(f"warning: only {len(rows)} labeled rows (README recommends 30-50)")

    with open(args.out, "w", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")
    print(f"wrote {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
